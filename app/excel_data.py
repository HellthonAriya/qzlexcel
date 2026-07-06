import re
from dataclasses import dataclass
from enum import Enum

import openpyxl

SHEET_EMOJIS = ["📗", "📘", "📙", "📕", "📔", "📒", "📓", "📖"]


def sheet_emoji(index: int) -> str:
    return SHEET_EMOJIS[index % len(SHEET_EMOJIS)]


class Status(Enum):
    EMPTY = "EMPTY"
    CHECK = "CHECK"
    CROSS = "CROSS"
    OTHER = "OTHER"


STATUS_ORDER = [Status.EMPTY, Status.CHECK, Status.CROSS, Status.OTHER]
STATUS_ICON = {
    Status.EMPTY: "⬜",
    Status.CHECK: "✅",
    Status.CROSS: "❌",
    Status.OTHER: "🔶",
}

PHONE_STATUS_TEXT = {
    Status.EMPTY: "",
    Status.CHECK: "جواب داد",
    Status.CROSS: "جواب نداد",
    Status.OTHER: "سایر",
}
ATTEND_STATUS_TEXT = {
    Status.EMPTY: "",
    Status.CHECK: "می‌آید",
    Status.CROSS: "نمی‌آید",
    Status.OTHER: "سایر",
}


def next_status(status: Status) -> Status:
    idx = STATUS_ORDER.index(status)
    return STATUS_ORDER[(idx + 1) % len(STATUS_ORDER)]


def parse_phone_status(text) -> Status:
    if text is None or not str(text).strip():
        return Status.EMPTY
    t = str(text).strip()
    if "نداد" in t:
        return Status.CROSS
    if "داد" in t:
        return Status.CHECK
    return Status.OTHER


def parse_attend_status(text) -> Status:
    if text is None or not str(text).strip():
        return Status.EMPTY
    t = re.sub(r"[\s‌]", "", str(text).strip())
    if "نمی" in t:
        return Status.CROSS
    if "می" in t:
        return Status.CHECK
    return Status.OTHER


def format_phone(value):
    if value is None or value == "":
        return None
    if isinstance(value, float):
        value = int(value)
    digits = re.sub(r"\D", "", str(value).strip())
    if not digits:
        return None
    return digits[-10:]


def _normalize_header(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip())


_HEADER_RULES = [
    ("redif", lambda h: h == "ردیف"),
    ("operator", lambda h: "اپراتور" in h),
    ("full_name", lambda h: "نام" in h and "خانوادگی" in h),
    ("grade", lambda h: h == "مقطع"),
    ("major", lambda h: h == "رشته"),
    ("service_type", lambda h: "نوع خدمات" in h),
    ("home_phone", lambda h: "تلفن منزل" in h),
    ("student_phone", lambda h: "دانش" in h and "آموز" in h),
    ("mother_phone", lambda h: "مادر" in h),
    ("father_phone", lambda h: "پدر" in h),
    ("call_date", lambda h: "تاریخ تماس" in h),
    ("phone_status", lambda h: "پاسخگویی" in h),
    ("attendance_status", lambda h: "حضور" in h),
    ("notes", lambda h: "توضیحات" in h),
]


def _build_col_map(ws, operator_col_letter=None):
    col_map = {}
    for cell in ws[1]:
        header = _normalize_header(cell.value)
        if not header:
            continue
        for key, predicate in _HEADER_RULES:
            if key in col_map:
                continue
            if predicate(header):
                col_map[key] = cell.column
                break
    if operator_col_letter:
        try:
            col_map["operator"] = openpyxl.utils.column_index_from_string(operator_col_letter.upper())
        except ValueError:
            pass
    return col_map


def _is_yellow(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb":
        return False
    rgb = str(fg.rgb or "").upper()
    return rgb.endswith("FFFF00")


@dataclass
class Record:
    sheet_key: str
    row: int
    redif: object
    operator: object
    full_name: object
    grade: object
    major: object
    service_type: object
    home_phone: object
    student_phone: object
    mother_phone: object
    father_phone: object
    call_date: object
    notes: object
    highlight_student: bool
    highlight_mother: bool
    highlight_father: bool
    hidden: bool
    phone_status: Status
    attendance_status: Status


def load_workbook_records(path, operator_col_letter=None):
    """Reads every worksheet that looks like a student list (has a
    recognizable name/ردیف column), whatever its name happens to be.
    Returns (records_by_sheet, sheet_labels) where sheet_labels maps the
    generated short key (s0, s1, ...) to the sheet's actual name, in the
    order the sheets appear in the workbook.

    operator_col_letter, if given (e.g. "A"), forces the اپراتور/admin
    column to that letter on every sheet instead of relying on the
    header text, for files where that column's header is missing or
    doesn't say "اپراتور".
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    records_by_sheet = {}
    sheet_labels = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_map = _build_col_map(ws, operator_col_letter)
        if "full_name" not in col_map and "redif" not in col_map:
            continue

        key = f"s{len(sheet_labels)}"
        records = {}
        for r in range(2, ws.max_row + 1):
            def val(field_key):
                col = col_map.get(field_key)
                return ws.cell(r, col).value if col else None

            full_name = val("full_name")
            redif = val("redif")
            if not full_name and not redif:
                continue

            def yellow(field_key):
                col = col_map.get(field_key)
                return bool(col) and _is_yellow(ws.cell(r, col))

            row_dim = ws.row_dimensions.get(r)
            hidden = bool(row_dim.hidden) if row_dim is not None else False

            operator = _normalize_header(val("operator")) or None

            records[r] = Record(
                sheet_key=key,
                row=r,
                redif=redif,
                operator=operator,
                full_name=full_name,
                grade=val("grade"),
                major=val("major"),
                service_type=val("service_type"),
                home_phone=val("home_phone"),
                student_phone=val("student_phone"),
                mother_phone=val("mother_phone"),
                father_phone=val("father_phone"),
                call_date=val("call_date"),
                notes=val("notes"),
                highlight_student=yellow("student_phone"),
                highlight_mother=yellow("mother_phone"),
                highlight_father=yellow("father_phone"),
                hidden=hidden,
                phone_status=parse_phone_status(val("phone_status")),
                attendance_status=parse_attend_status(val("attendance_status")),
            )
        records_by_sheet[key] = records
        sheet_labels[key] = sheet_name
    return records_by_sheet, sheet_labels


def export_workbook(source_path, records_by_sheet, sheet_labels, out_path):
    wb = openpyxl.load_workbook(source_path)
    for key, records in records_by_sheet.items():
        sheet_name = sheet_labels.get(key)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_map = _build_col_map(ws)
        phone_col = col_map.get("phone_status")
        attend_col = col_map.get("attendance_status")
        for rec in records.values():
            if phone_col:
                ws.cell(rec.row, phone_col).value = PHONE_STATUS_TEXT[rec.phone_status] or None
            if attend_col:
                ws.cell(rec.row, attend_col).value = ATTEND_STATUS_TEXT[rec.attendance_status] or None
    wb.save(out_path)
    return out_path
